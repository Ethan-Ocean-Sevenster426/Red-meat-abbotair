"""Extract all data from MySQL into seed-data/*.xlsx for seeding on a fresh setup."""
import os, sys, django
os.environ['DJANGO_SETTINGS_MODULE'] = 'rmaa_backend.settings'
sys.path.insert(0, os.path.dirname(__file__))
os.chdir(os.path.dirname(__file__))
django.setup()

from pathlib import Path
from django.db import connection
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

TABLES = [
    'Users',
    'AbattoirMaster',
    'TransformationMaster',
    'GovernmentMaster',
    'IndustryMaster',
    'AssociatedMembersMaster',
    'STTTrainingReport',
    'TrainingReport',
    'ResidueMonitoring',
    'CustomAbattoirs',
    'Learners',
    'Facilitators',
    'FeeStructure',
    'AuditLog',
    'UserColumnPreferences',
    'Invitations',
]

out_dir = Path(__file__).parent.parent / 'seed-data'
out_dir.mkdir(exist_ok=True)

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(start_color='0078D4', end_color='0078D4', fill_type='solid')
header_align = Alignment(vertical='center', horizontal='center')

summary = []

for table in TABLES:
    try:
        with connection.cursor() as c:
            c.execute(f'SELECT COUNT(*) FROM {table}')
            count = c.fetchone()[0]

        if count == 0:
            print(f'  {table}: 0 rows — skipped')
            summary.append((table, 0, 'empty'))
            continue

        with connection.cursor() as c:
            c.execute(f'SELECT * FROM {table}')
            columns = [col[0] for col in c.description]
            rows = c.fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = table

        # Header
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align

        # Data
        for row in rows:
            ws.append([str(v) if v is not None else '' for v in row])

        # Auto-width
        for i, col_name in enumerate(columns, 1):
            max_len = len(col_name)
            for row in rows[:100]:
                val = row[i - 1]
                if val is not None:
                    max_len = max(max_len, min(len(str(val)), 50))
            ws.column_dimensions[ws.cell(1, i).column_letter].width = max_len + 2

        # Freeze header
        ws.freeze_panes = 'A2'

        path = out_dir / f'{table}.xlsx'
        wb.save(path)
        print(f'  {table}: {count} rows -> {table}.xlsx')
        summary.append((table, count, 'ok'))

    except Exception as e:
        print(f'  {table}: ERROR — {e}')
        summary.append((table, 0, f'error: {e}'))

print('\n' + '=' * 60)
print('EXTRACTION SUMMARY')
print('=' * 60)
total = 0
for table, count, status in summary:
    print(f'  {table:30s} {count:>8}  {status}')
    total += count
print('-' * 60)
print(f'  {"TOTAL":30s} {total:>8}')
print(f'\nFiles saved to: {out_dir}')
