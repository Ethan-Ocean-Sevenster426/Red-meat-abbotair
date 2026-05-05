"""API views for the RMAA Django backend.

Mirrors the behavior of server/routes/*.js — URL paths, request/response shapes,
and query-param semantics are preserved so the existing React frontend works
unchanged.
"""
from __future__ import annotations
import base64
import hashlib
import json
import re
import uuid
from datetime import datetime, timedelta
from io import BytesIO
from pathlib import Path

from django.conf import settings
from django.db import connection
from django.http import FileResponse
from django.views.decorators.clickjacking import xframe_options_exempt
from django.utils import timezone
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response

from . import email as email_svc
from .column_maps import (
    ABATTOIR_COLS, ABATTOIR_BRACKETS,
    TRANSFORMATION_COLS, TRANSFORMATION_BRACKETS,
    GOVERNMENT_COLS, GOVERNMENT_BRACKETS,
    INDUSTRY_COLS, INDUSTRY_BRACKETS,
    ASSOCIATED_COLS, ASSOCIATED_BRACKETS,
    STT_COLS, STT_BRACKETS,
    TRAINING_COLS, TRAINING_BRACKETS,
    LEARNER_COLS, LEARNER_BRACKETS,
    FEE_STRUCTURE_COLS, FEE_STRUCTURE_BRACKETS,
    RESIDUE_ALL_COLS, RESIDUE_INSERT_COLS, RESIDUE_HEADER_MAP,
    AUDIT_NAME_JOIN,
)
from .helpers import (
    append_audit_log, apply_column_filters, build_master_crud,
    rows_to_dicts, safe_col_ref, sanitize_fs_name,
)
from .models import (
    AbattoirMaster, TransformationMaster, GovernmentMaster, IndustryMaster,
    AssociatedMembersMaster, STTTrainingReport, TrainingReport,
    CustomAbattoir, Facilitator, FeeStructure, Learner, User, Invitation, UserColumnPreferences,
)


# ===========================================================================
#  status + health
# ===========================================================================
@api_view(['GET'])
def status_view(request):
    return Response({'status': 'ok'})


@api_view(['GET'])
def health_view(request):
    checks = []
    try:
        with connection.cursor() as c:
            c.execute('SELECT 1')
            c.fetchone()
        checks.append({'check': 'DB connection', 'ok': True, 'detail': 'OK'})
        for table in ['Users', 'ResidueMonitoringTemp', 'ResidueMonitoring']:
            with connection.cursor() as c:
                c.execute(f'SELECT COUNT(*) FROM {table}')
                n = c.fetchone()[0]
            checks.append({'check': f'Table {table}', 'ok': True, 'detail': f'{n} rows'})
    except Exception as e:
        checks.append({'check': 'DB test', 'ok': False, 'detail': str(e)})
    all_ok = all(c['ok'] for c in checks)
    return Response({'ok': all_ok, 'checks': checks}, status=200 if all_ok else 500)


# ===========================================================================
#  auth
# ===========================================================================
@api_view(['POST'])
def login_view(request):
    email = (request.data.get('email') or '').strip()
    password = request.data.get('password') or ''
    if not email or not password:
        return Response({'message': 'Email and password are required.'}, status=400)
    try:
        user = User.objects.filter(email=email, password=password).first()
        if not user:
            return Response({'message': 'Invalid email or password.'}, status=401)
        return Response({
            'user': {
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'role': user.role,
                'displayName': user.display_name,
            }
        })
    except Exception as e:
        return Response({'message': f'Database error during login: {e}'}, status=500)


# ===========================================================================
#  users + invitations
# ===========================================================================
@api_view(['GET'])
def users_list_view(request):
    rows = []
    for u in User.objects.order_by('id'):
        rows.append({
            'id': u.id, 'username': u.username, 'email': u.email,
            'displayName': u.display_name, 'role': u.role,
            'permissions': u.permissions, 'created_at': u.created_at,
        })
    return Response({'users': rows})


@api_view(['POST'])
def users_invite_view(request):
    email = (request.data.get('email') or '').strip()
    display_name = (request.data.get('displayName') or '').strip()
    invited_by = (request.data.get('invitedBy') or 'Admin').strip()
    if not email or not display_name:
        return Response({'message': 'Email and name are required.'}, status=400)

    if User.objects.filter(email=email).exists():
        return Response({'message': 'A user with this email already exists.'}, status=409)

    user = User.objects.create(
        username=email, email=email, display_name=display_name, role='user', password=''
    )
    token = str(uuid.uuid4())
    expires = (datetime.utcnow() + timedelta(days=7)).isoformat()
    Invitation.objects.create(
        token=token, email=email, user_id=user.id,
        invited_by=invited_by, expires_at=expires,
    )

    invite_url = f"{settings.APP_BASE_URL}/accept-invite?token={token}"
    email_result = email_svc.send_invite_email(
        to=email, invited_by=invited_by, invite_url=invite_url
    )
    return Response({
        'ok': True, 'inviteUrl': invite_url,
        'emailSent': email_result.get('ok', False), 'userId': user.id,
    })


@api_view(['GET'])
def users_invite_lookup_view(request, token):
    inv = Invitation.objects.filter(token=token, accepted=False).first()
    if not inv:
        return Response({'message': 'Invitation not found or already used.'}, status=404)
    try:
        exp = datetime.fromisoformat((inv.expires_at or '').replace('Z', ''))
        if exp < datetime.utcnow():
            return Response({'message': 'This invitation has expired.'}, status=410)
    except ValueError:
        pass
    user = User.objects.filter(id=inv.user_id).first()
    return Response({
        'valid': True, 'email': inv.email,
        'displayName': user.display_name if user else '',
        'invitedBy': inv.invited_by,
    })


@api_view(['POST'])
def users_invite_accept_view(request):
    token = request.data.get('token')
    password = request.data.get('password')
    display_name = (request.data.get('displayName') or '').strip()
    if not token or not password:
        return Response({'message': 'Token and password required.'}, status=400)
    inv = Invitation.objects.filter(token=token, accepted=False).first()
    if not inv:
        return Response({'message': 'Invalid or already used invitation.'}, status=404)
    try:
        exp = datetime.fromisoformat((inv.expires_at or '').replace('Z', ''))
        if exp < datetime.utcnow():
            return Response({'message': 'Invitation has expired.'}, status=410)
    except ValueError:
        pass

    user = User.objects.filter(id=inv.user_id).first()
    if not user:
        return Response({'message': 'User not found.'}, status=404)
    user.password = password
    if display_name:
        user.display_name = display_name
    user.save()

    inv.accepted = True
    inv.accepted_at = timezone.now()
    inv.save()
    return Response({'ok': True})


@api_view(['PUT', 'PATCH'])
def users_permissions_view(request, pk):
    user = User.objects.filter(id=pk).first()
    if not user:
        return Response({'message': 'User not found'}, status=404)
    user.permissions = json.dumps(request.data.get('permissions') or {})
    user.save()
    return Response({'ok': True})


@api_view(['PUT', 'PATCH'])
def users_role_view(request, pk):
    user = User.objects.filter(id=pk).first()
    if not user:
        return Response({'message': 'User not found'}, status=404)
    user.role = request.data.get('role') or 'user'
    user.save()
    return Response({'ok': True})


@api_view(['DELETE'])
def users_delete_view(request, pk):
    Invitation.objects.filter(user_id=pk).delete()
    User.objects.filter(id=pk).delete()
    return Response({'ok': True})


# ===========================================================================
#  user column preferences
# ===========================================================================
@api_view(['GET', 'PUT'])
def user_prefs_view(request):
    try:
        user_id = int(request.query_params.get('userId') or 0)
    except ValueError:
        user_id = 0
    page = (request.query_params.get('page') or '').strip()

    if request.method == 'GET':
        if not user_id or not page:
            return Response({'hiddenColumns': [], 'columnOrder': []})
        pref = UserColumnPreferences.objects.filter(user_id=user_id, page_name=page).first()
        if not pref:
            return Response({'hiddenColumns': [], 'columnOrder': []})
        return Response({
            'hiddenColumns': json.loads(pref.hidden_columns or '[]'),
            'columnOrder': json.loads(pref.column_order or '[]'),
        })

    if not user_id or not page:
        return Response({'message': 'Missing userId or page'}, status=400)
    hidden = json.dumps(request.data.get('hiddenColumns') or [])
    order = json.dumps(request.data.get('columnOrder') or [])
    UserColumnPreferences.objects.update_or_create(
        user_id=user_id, page_name=page,
        defaults={'hidden_columns': hidden, 'column_order': order},
    )
    return Response({'ok': True})


# ===========================================================================
#  audit log
# ===========================================================================
@api_view(['GET'])
def audit_log_view(request):
    try:
        page = max(1, int(request.query_params.get('page') or '1'))
        page_size = min(200, int(request.query_params.get('size') or '50'))
    except ValueError:
        page, page_size = 1, 50
    offset = (page - 1) * page_size

    table = request.query_params.get('table') or ''
    user_filter = request.query_params.get('user') or ''
    year = request.query_params.get('year') or ''
    month = request.query_params.get('month') or ''
    days = request.query_params.get('days') or ''

    where: list[str] = []
    params: list = []
    vendor = connection.vendor

    if table:
        where.append('a.table_name = %s')
        params.append(table)
    if user_filter:
        where.append('a.modified_by LIKE %s')
        params.append(f'%{user_filter}%')
    if year:
        yrs = [int(y) for y in year.split(',') if y.strip().isdigit()]
        yrs = [y for y in yrs if y > 0]
        if yrs:
            ph = ','.join(['%s'] * len(yrs))
            if vendor == 'sqlite':
                where.append(f"CAST(strftime('%Y', a.created_at) AS INTEGER) IN ({ph})")
            else:
                where.append(f"YEAR(a.created_at) IN ({ph})")
            params.extend(yrs)
    if month:
        mos = [int(m) for m in month.split(',') if m.strip().isdigit()]
        mos = [m for m in mos if 1 <= m <= 12]
        if mos:
            ph = ','.join(['%s'] * len(mos))
            if vendor == 'sqlite':
                where.append(f"CAST(strftime('%m', a.created_at) AS INTEGER) IN ({ph})")
            else:
                where.append(f"MONTH(a.created_at) IN ({ph})")
            params.extend(mos)
    if days:
        try:
            d = int(days)
            if d > 0:
                where.append("a.created_at >= %s")
                params.append(datetime.utcnow() - timedelta(days=d))
        except ValueError:
            pass

    where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''

    name_join = ''
    name_select = ', NULL AS record_name'
    if table in AUDIT_NAME_JOIN:
        tbl, col = AUDIT_NAME_JOIN[table]
        name_join = f'LEFT JOIN {tbl} AS src ON src.id = a.record_id'
        name_select = f', src.{col} AS record_name'

    with connection.cursor() as c:
        c.execute(f'SELECT COUNT(*) FROM AuditLog a {name_join} {where_sql}', params)
        total = c.fetchone()[0]
        c.execute(
            f"""SELECT a.id, a.table_name, a.record_id, a.action_type {name_select},
                       a.modified_by, a.modified_time, a.modified_fields,
                       a.old_values, a.new_values, a.created_at
                FROM AuditLog a {name_join} {where_sql}
                ORDER BY a.created_at DESC
                LIMIT %s OFFSET %s""",
            params + [page_size, offset],
        )
        rows = rows_to_dicts(c)

        if table:
            c.execute(
                """SELECT DISTINCT modified_by FROM AuditLog
                   WHERE table_name = %s AND modified_by IS NOT NULL AND modified_by <> ''
                   ORDER BY modified_by""",
                [table],
            )
        else:
            c.execute(
                """SELECT DISTINCT modified_by FROM AuditLog
                   WHERE modified_by IS NOT NULL AND modified_by <> ''
                   ORDER BY modified_by"""
            )
        users = [r[0] for r in c.fetchall()]

    name_label = None
    if table in AUDIT_NAME_JOIN:
        col = AUDIT_NAME_JOIN[table][1]
        name_label = ' '.join(p.capitalize() for p in col.split('_'))

    return Response({
        'total': total, 'page': page, 'pageSize': page_size,
        'rows': rows, 'users': users, 'nameLabel': name_label,
    })


# ===========================================================================
#  master CRUD resources (via generic builder)
# ===========================================================================
abattoir_views = build_master_crud(
    AbattoirMaster, columns=ABATTOIR_COLS, bracket_cols=ABATTOIR_BRACKETS,
    audit_table='AbattoirMaster',
)
transformation_views = build_master_crud(
    TransformationMaster, columns=TRANSFORMATION_COLS, bracket_cols=TRANSFORMATION_BRACKETS,
    audit_table='TransformationMaster',
)
government_views = build_master_crud(
    GovernmentMaster, columns=GOVERNMENT_COLS, bracket_cols=GOVERNMENT_BRACKETS,
    audit_table='GovernmentMaster',
)
industry_views = build_master_crud(
    IndustryMaster, columns=INDUSTRY_COLS, bracket_cols=INDUSTRY_BRACKETS,
    audit_table='IndustryMaster',
)
associated_views = build_master_crud(
    AssociatedMembersMaster, columns=ASSOCIATED_COLS, bracket_cols=ASSOCIATED_BRACKETS,
    audit_table='AssociatedMembersMaster',
)
training_views = build_master_crud(
    TrainingReport, columns=TRAINING_COLS, bracket_cols=TRAINING_BRACKETS,
    audit_table='TrainingReport',
)
stt_views = build_master_crud(
    STTTrainingReport, columns=STT_COLS, bracket_cols=STT_BRACKETS,
    audit_table='STTTrainingReport',
)

learner_views = build_master_crud(
    Learner, columns=LEARNER_COLS, bracket_cols=LEARNER_BRACKETS,
    audit_table='Learners',
)

fee_structure_views = build_master_crud(
    FeeStructure, columns=FEE_STRUCTURE_COLS, bracket_cols=FEE_STRUCTURE_BRACKETS,
    audit_table='FeeStructure',
)


# ===========================================================================
#  abattoir extras (names, custom, send form)
# ===========================================================================
@api_view(['GET'])
def abattoir_names_view(request):
    with connection.cursor() as c:
        c.execute(
            """SELECT abattoir_name AS name, municipality, province, lh AS thru_put
               FROM AbattoirMaster
               WHERE abattoir_name IS NOT NULL AND abattoir_name <> ''
               ORDER BY abattoir_name"""
        )
        registered = rows_to_dicts(c)
        c.execute("SELECT id, name, NULL AS municipality FROM CustomAbattoirs ORDER BY name")
        custom = rows_to_dicts(c)
    return Response({'registered': registered, 'custom': custom})


@api_view(['POST'])
def abattoir_custom_add_view(request):
    name = (request.data.get('name') or '').strip()
    if not name:
        return Response({'message': 'Name required'}, status=400)
    row = CustomAbattoir.objects.create(name=name)
    return Response({'id': row.id, 'name': row.name})


@api_view(['DELETE'])
def abattoir_custom_delete_view(request, pk):
    CustomAbattoir.objects.filter(id=pk).delete()
    return Response({'ok': True})


@api_view(['POST'])
def abattoir_send_form_view(request, pk):
    from docxtpl import DocxTemplate

    row = AbattoirMaster.objects.filter(id=pk).values().first()
    if not row:
        return Response({'message': 'Record not found.'}, status=404)

    def v(x):
        return str(x) if x is not None and x != '' else ''

    data = {
        'AbattoirName': v(row['abattoir_name']),
        'RegistrationNr': v(row['rc_nr']),
        'ExportNumber': v(row['za_nr']),
        'Units': v(row['units']),
        'Slaughter': v(row['amount_slaughtered']),
        'Throughput': v(row['lh']),
        'Cattle': v(row['cattle']), 'Sheep': v(row['sheep']), 'Pigs': v(row['pig']),
        'Game': v(row['game']), 'Horses': v(row['horses']),
        'VATNumber': v(row['vat_number']), 'Halaal': v(row['halaal']), 'Kosher': v(row['kosher']),
        'SETA': v(row['seta']),
        'Telephone1': v(row['tel_1']), 'Telephone2': v(row['tel_2']), 'FAX': v(row['fax']),
        'Municipality': v(row['municipality']), 'PostalAddress': v(row['postal_address']),
        'Owner': v(row['owner']), 'OwnerEmail': v(row['owner_email']), 'OwnerCell': v(row['owner_cell']),
        'Manager': v(row['manager']), 'ManagerEmail': v(row['manager_email']), 'ManagerCell': v(row['manager_cell']),
        'Training': v(row['training']), 'TrainingEmail': v(row['training_email']), 'TrainingCell': v(row['training_cell']),
        'Accounts': v(row['accounts']), 'AccountsEmail': v(row['accounts_email']), 'AccountsCell': v(row['accounts_cell']),
        'MeatInspectionServices': v(row['assignee_name']),
        'MIServiceContactDetails': v(row['assignee_contact_name']) + (
            f" / {row['assignee_contact_number']}" if row.get('assignee_contact_number') else ''),
        'MeatInspectors': v(row['meat_inspectors']), 'MeatExaminers': v(row['meat_examiner']),
        'Graders': v(row['grader']), 'Classifier': v(row['classification']),
        'QAManager': v(row['qa_manager']), 'FloorSupervisor': v(row['floor_supervisor']),
        'PhysicalAddress': v(row['physical_address']),
        'GPSCoordinates': ' / '.join([x for x in [row.get('gps_1'), row.get('gps_2')] if x]),
    }

    template_path = Path(settings.PROJECT_ROOT) / 'RMAA Database Form.docx'
    if not template_path.exists():
        return Response({'message': f'Template not found: {template_path}'}, status=500)

    try:
        from docx import Document
        from copy import deepcopy
        from lxml import etree

        doc = Document(str(template_path))
        body = doc.element.body
        W = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'
        ns = {'w': W}

        # Fill Word content controls (SDT) — match by alias or tag
        for sdt in body.iter(f'{{{W}}}sdt'):
            tag = sdt.find(f'.//{{{W}}}tag', ns)
            alias = sdt.find(f'.//{{{W}}}alias', ns)
            tag_val = tag.get(f'{{{W}}}val') if tag is not None else None
            alias_val = alias.get(f'{{{W}}}val') if alias is not None else None
            # Prefer alias — template has at least one SDT with mismatched tag
            key = alias_val if alias_val in data else tag_val
            if not key or key not in data:
                continue
            content = sdt.find(f'{{{W}}}sdtContent', ns)
            if content is None:
                continue
            # Find first run, clear all <w:t> in this content, set first to value
            runs = content.findall(f'.//{{{W}}}r', ns)
            if not runs:
                continue
            first_set = False
            for r in runs:
                for t in r.findall(f'{{{W}}}t', ns):
                    if not first_set:
                        t.text = data[key]
                        t.set('{http://www.w3.org/XML/1998/namespace}space', 'preserve')
                        first_set = True
                    else:
                        t.text = ''

        buf = BytesIO()
        doc.save(buf)
        doc_bytes = buf.getvalue()
    except Exception as e:
        return Response({'message': f'Failed to build form: {e}'}, status=500)

    filename = f"RMAA Database Form - {row['abattoir_name'] or 'Abattoir'}.docx"
    recipient = (request.data.get('to') if request.data else None) or 'training@rmaa.co.za'
    result = email_svc.send_database_form(
        to=recipient,
        abattoir_name=row['abattoir_name'] or 'Unknown Abattoir',
        training_email=row.get('training_email') or '',
        doc_buffer=doc_bytes,
        filename=filename,
    )
    if not result.get('ok'):
        return Response(
            {'message': f"Form generated but email failed: {result.get('reason')}"},
            status=500,
        )

    now = datetime.now()
    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
              'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    sent_date = f"{now.day:02d}-{months[now.month - 1]}-{str(now.year)[2:]}"

    AbattoirMaster.objects.filter(id=pk).update(date_mail_sent=sent_date)
    append_audit_log('AbattoirMaster', pk, {
        'modified_by': 'System',
        'modified_fields': json.dumps(['date_mail_sent']),
        'old_values': json.dumps({'date_mail_sent': row.get('date_mail_sent') or ''}),
        'new_values': json.dumps({'date_mail_sent': sent_date}),
    })

    return Response({
        'ok': True, 'dateSent': sent_date,
        'message': f"Database form sent successfully for {row['abattoir_name']}.",
    })


# ===========================================================================
#  STT Training Report extras — custom list filters and specialised endpoints
# ===========================================================================
@api_view(['GET', 'POST'])
def stt_list_create_view(request):
    if request.method == 'POST':
        return stt_views['list_create'](request._request)

    try:
        page = max(1, int(request.query_params.get('page') or '1'))
        page_size = min(100, int(request.query_params.get('size') or '50'))
    except ValueError:
        page, page_size = 1, 50
    offset = (page - 1) * page_size

    reserved = {'page', 'size', 'sortCol', 'sortDir',
                '_years', '_months', '_provinces', '_idCheck'}
    where_parts, params = apply_column_filters(
        request.query_params.dict(), STT_COLS, STT_BRACKETS, reserved,
    )

    vendor = connection.vendor

    def yfn(col):
        return (f"CAST(strftime('%Y', {col}) AS INTEGER)"
                if vendor == 'sqlite' else f"YEAR({col})")

    def mfn(col):
        return (f"CAST(strftime('%m', {col}) AS INTEGER)"
                if vendor == 'sqlite' else f"MONTH({col})")

    years_raw = request.query_params.get('_years') or ''
    if years_raw:
        yrs = [int(y) for y in years_raw.split(',') if y.strip().isdigit()]
        yrs = [y for y in yrs if 0 < y < 3000]
        if yrs:
            ph = ','.join(['%s'] * len(yrs))
            where_parts.append(f"{yfn('training_start_date')} IN ({ph})")
            params.extend(yrs)

    months_raw = request.query_params.get('_months') or ''
    if months_raw:
        mos = [int(m) for m in months_raw.split(',') if m.strip().isdigit()]
        mos = [m for m in mos if 1 <= m <= 12]
        if mos:
            ph = ','.join(['%s'] * len(mos))
            where_parts.append(f"{mfn('training_start_date')} IN ({ph})")
            params.extend(mos)

    provs_raw = request.query_params.get('_provinces') or ''
    if provs_raw:
        provs = [p for p in provs_raw.split(',') if p]
        if provs:
            ph = ','.join(['%s'] * len(provs))
            where_parts.append(f"province IN ({ph})")
            params.extend(provs)

    id_check = request.query_params.get('_idCheck') or ''
    if id_check == 'duplicate':
        where_parts.append(
            "id_number IN (SELECT id_number FROM STTTrainingReport "
            "WHERE id_number IS NOT NULL AND id_number <> '' "
            "GROUP BY id_number HAVING COUNT(*) > 1)"
        )
    elif id_check == 'incorrect':
        length_fn = 'LENGTH' if vendor == 'sqlite' else 'CHAR_LENGTH'
        where_parts.append(
            f"id_number IS NOT NULL AND id_number <> '' AND {length_fn}(id_number) <> 13"
        )
    elif id_check == 'missing':
        where_parts.append("(id_number IS NULL OR id_number = '')")

    where_sql = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    sort_col_raw = request.query_params.get('sortCol') or ''
    sort_col = safe_col_ref(sort_col_raw, STT_BRACKETS) if sort_col_raw in STT_COLS else 'id'
    sort_dir = 'DESC' if request.query_params.get('sortDir') == 'desc' else 'ASC'

    with connection.cursor() as c:
        c.execute(f'SELECT COUNT(*) FROM STTTrainingReport {where_sql}', params)
        total = c.fetchone()[0]
        c.execute(
            f'SELECT * FROM STTTrainingReport {where_sql} ORDER BY {sort_col} {sort_dir} '
            f'LIMIT %s OFFSET %s',
            params + [page_size, offset],
        )
        rows = rows_to_dicts(c)
    return Response({'total': total, 'page': page, 'pageSize': page_size, 'rows': rows})


@api_view(['GET', 'POST'])
def learner_list_create_view(request):
    """Custom list for Learners with _idCheck support."""
    if request.method == 'POST':
        return learner_views['list_create'](request._request)

    try:
        page = max(1, int(request.query_params.get('page') or '1'))
        page_size = min(200, int(request.query_params.get('size') or '50'))
    except ValueError:
        page, page_size = 1, 50
    offset = (page - 1) * page_size

    vendor = connection.vendor
    length_fn = 'LENGTH' if vendor == 'sqlite' else 'CHAR_LENGTH'
    reserved = {'page', 'size', 'sortCol', 'sortDir', '_idCheck'}
    where_parts, params = apply_column_filters(
        request.query_params.dict(), LEARNER_COLS, LEARNER_BRACKETS, reserved,
    )

    id_check = request.query_params.get('_idCheck') or ''
    if id_check == 'duplicate':
        where_parts.append(
            "id_number IN (SELECT id_number FROM Learners "
            "WHERE id_number IS NOT NULL AND id_number <> '' "
            "GROUP BY id_number HAVING COUNT(*) > 1)"
        )
    elif id_check == 'incorrect':
        where_parts.append(
            f"id_number IS NOT NULL AND id_number <> '' AND {length_fn}(id_number) <> 13"
        )
    elif id_check == 'missing':
        where_parts.append("(id_number IS NULL OR id_number = '')")

    where_sql = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    sort_col_raw = request.query_params.get('sortCol') or ''
    sort_col = sort_col_raw if sort_col_raw in LEARNER_COLS else 'id'
    sort_dir = 'DESC' if request.query_params.get('sortDir') == 'desc' else 'ASC'

    with connection.cursor() as c:
        c.execute(f'SELECT COUNT(*) FROM Learners {where_sql}', params)
        total = c.fetchone()[0]
        c.execute(
            f'SELECT * FROM Learners {where_sql} ORDER BY {sort_col} {sort_dir} '
            f'LIMIT %s OFFSET %s',
            params + [page_size, offset],
        )
        rows = rows_to_dicts(c)
    return Response({'total': total, 'page': page, 'pageSize': page_size, 'rows': rows})


@api_view(['POST'])
def learner_merge_view(request):
    """Merge two learners: keep primary, combine work_stations, delete secondary."""
    primary_id = request.data.get('primary_id')
    secondary_id = request.data.get('secondary_id')
    merged_by = request.data.get('modified_by') or ''

    if not primary_id or not secondary_id or primary_id == secondary_id:
        return Response({'message': 'Two different learner IDs are required.'}, status=400)

    with connection.cursor() as c:
        c.execute('SELECT * FROM Learners WHERE id = %s', [primary_id])
        primary = rows_to_dicts(c)
        c.execute('SELECT * FROM Learners WHERE id = %s', [secondary_id])
        secondary = rows_to_dicts(c)

    if not primary or not secondary:
        return Response({'message': 'One or both learners not found.'}, status=404)

    primary = primary[0]
    secondary = secondary[0]

    # Combine work stations
    ws_primary = set(filter(None, (primary.get('work_stations') or '').split(', ')))
    ws_secondary = set(filter(None, (secondary.get('work_stations') or '').split(', ')))
    combined_ws = ', '.join(sorted(ws_primary | ws_secondary))

    now = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')

    # Update primary with combined work stations
    with connection.cursor() as c:
        c.execute(
            'UPDATE Learners SET work_stations = %s, modified_by = %s, modified_time = %s WHERE id = %s',
            [combined_ws, merged_by, now, primary_id],
        )

    # Log the merge on the primary record
    append_audit_log('Learners', int(primary_id), {
        'modified_by': merged_by,
        'modified_time': now,
        'modified_fields': f'Merged with learner #{secondary_id}',
        'old_values': json.dumps({
            'work_stations': primary.get('work_stations') or '',
        }),
        'new_values': json.dumps({
            'work_stations': combined_ws,
            'merged_from': f"#{secondary_id} {secondary.get('name') or ''} {secondary.get('surname') or ''} ({secondary.get('id_number') or 'no ID'})".strip(),
        }),
    }, action_type='EDIT')

    # Log the deletion of the secondary
    skip = {'id', 'modified_by', 'modified_time', 'modified_fields', 'old_values', 'new_values'}
    old_data = {k: str(v) for k, v in secondary.items() if k not in skip and v is not None and str(v).strip()}
    append_audit_log('Learners', int(secondary_id), {
        'modified_by': merged_by,
        'modified_time': now,
        'modified_fields': ','.join(old_data.keys()) if old_data else 'merged into #' + str(primary_id),
        'old_values': json.dumps(old_data) if old_data else '',
        'new_values': f'Merged into learner #{primary_id}',
    }, action_type='DELETE')

    # Delete secondary
    with connection.cursor() as c:
        c.execute('DELETE FROM Learners WHERE id = %s', [secondary_id])

    return Response({'ok': True, 'work_stations': combined_ws})


@api_view(['GET'])
def stt_learner_summary_view(request):
    """Consolidated learner view — one row per person, work stations aggregated."""
    try:
        page = max(1, int(request.query_params.get('page') or '1'))
        page_size = min(200, int(request.query_params.get('size') or '50'))
    except ValueError:
        page, page_size = 1, 50
    offset = (page - 1) * page_size

    vendor = connection.vendor
    length_fn = 'LENGTH' if vendor == 'sqlite' else 'CHAR_LENGTH'
    group_concat = 'GROUP_CONCAT(DISTINCT work_station)' if vendor == 'sqlite' else "GROUP_CONCAT(DISTINCT work_station SEPARATOR ', ')"

    where_parts = []
    params = []

    # Column filters
    filterable = {'name', 'surname', 'id_number', 'year_of_birth', 'age', 'citizen', 'race_gender'}
    reserved = {'page', 'size', 'sortCol', 'sortDir', '_idCheck'}
    for key, val in request.query_params.dict().items():
        if key in reserved or key not in filterable:
            continue
        v = (val or '').strip()
        if not v:
            continue
        if v == '__blank__':
            where_parts.append(f"({key} = '' OR {key} IS NULL)")
        else:
            where_parts.append(f"{key} LIKE %s")
            params.append(f'%{v}%')

    # ID check (no duplicate — only incorrect and missing)
    id_check = request.query_params.get('_idCheck') or ''
    if id_check == 'incorrect':
        where_parts.append(
            f"id_number IS NOT NULL AND id_number <> '' AND {length_fn}(id_number) <> 13"
        )
    elif id_check == 'missing':
        where_parts.append("(id_number IS NULL OR id_number = '')")

    where_sql = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    # Sort
    sort_col_raw = request.query_params.get('sortCol') or ''
    valid_sort = {'name', 'surname', 'id_number', 'year_of_birth', 'age', 'citizen', 'race_gender', 'work_stations', 'times_trained'}
    sort_col = sort_col_raw if sort_col_raw in valid_sort else 'surname'
    sort_dir = 'DESC' if request.query_params.get('sortDir') == 'desc' else 'ASC'

    # Subquery to consolidate learners
    inner_sql = f"""
        SELECT
          name, surname, id_number,
          MAX(year_of_birth) AS year_of_birth,
          MAX(age) AS age,
          MAX(citizen) AS citizen,
          MAX(race_gender) AS race_gender,
          {group_concat} AS work_stations,
          COUNT(*) AS times_trained
        FROM STTTrainingReport
        {where_sql}
        GROUP BY name, surname, id_number
    """

    with connection.cursor() as c:
        # Total count
        c.execute(f"SELECT COUNT(*) FROM ({inner_sql}) AS sub", params)
        total = c.fetchone()[0]

        # Paginated data
        c.execute(
            f"{inner_sql} ORDER BY {sort_col} {sort_dir} LIMIT %s OFFSET %s",
            params + [page_size, offset],
        )
        rows = rows_to_dicts(c)

    return Response({'total': total, 'page': page, 'pageSize': page_size, 'rows': rows})


@api_view(['GET'])
def stt_learner_summary_count_view(request):
    vendor = connection.vendor
    with connection.cursor() as c:
        c.execute(
            "SELECT COUNT(*) FROM ("
            "  SELECT name, surname, id_number FROM STTTrainingReport"
            "  GROUP BY name, surname, id_number"
            ") AS sub"
        )
        n = c.fetchone()[0]
    return Response({'count': n})


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def stt_parse_excel_view(request):
    from openpyxl import load_workbook
    f = request.FILES.get('file')
    if not f:
        return Response({'message': 'No file uploaded.'}, status=400)
    try:
        wb = load_workbook(BytesIO(f.read()), data_only=True)
        ws = wb.worksheets[0]

        def cv(row_n, col_letter):
            """Read a cell, normalising Excel float/string artifacts.

            ID digit cells may come back as int, float (8.0), or even str ("8.0")
            depending on how the template was saved. All three must collapse to
            a single-digit string — otherwise joining 13 of them produces
            "8.06.00.08.0..." artefacts in the resulting id_number.
            """
            cell = ws[f'{col_letter}{row_n}']
            v = cell.value
            if v is None:
                return ''
            if isinstance(v, float) and v.is_integer():
                return str(int(v))
            s = str(v).strip()
            if re.match(r'^-?\d+\.0+$', s):
                return s.split('.')[0]
            return s

        session = {
            'programme': cv(16, 'C'),
            'specie': cv(16, 'G'),
            'facilitator': cv(16, 'L'),
            'contact': cv(16, 'Y'),
        }
        id_cols = ['L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X']
        rows_out = []
        for r in range(19, 49):
            work_station = cv(r, 'D')
            surname = cv(r, 'F')
            name = cv(r, 'I')
            race = cv(r, 'Z')
            gender = cv(r, 'AA')
            # For ID digit cells, always take the integer part only (cells
            # here are supposed to hold a single digit 0-9; a "8.06" cell is
            # treated as "8" and any stray non-digits stripped).
            id_digit_parts = []
            for c in id_cols:
                raw = cv(r, c)
                if '.' in raw:
                    raw = raw.split('.')[0]
                id_digit_parts.append(re.sub(r'\D', '', raw))
            id_number = ''.join(id_digit_parts)
            if not surname and not name and not id_number:
                continue
            rg = (race + gender).strip().upper()
            # Set corresponding race-gender column to "1" for this row
            rg_map = {
                'AM': 'am', 'AF': 'af', 'AD': 'ad',
                'CM': 'cm', 'CF': 'cf', 'CD': 'cd',
                'IM': 'im', 'IF': 'if_col', 'ID': 'id_2',
                'WM': 'wm', 'WF': 'wf', 'WD': 'wd',
            }
            rg_flags = {col: '' for col in rg_map.values()}
            if rg in rg_map:
                rg_flags[rg_map[rg]] = '1'
            rows_out.append({
                'work_station': work_station, 'surname': surname, 'name': name,
                'id_number': id_number, 'race_gender': rg,
                **rg_flags,
            })
        return Response({'session': session, 'rows': rows_out})
    except Exception as e:
        return Response({'message': f'Failed to parse Excel file: {e}'}, status=500)


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def stt_export_pdf_view(request):
    f = request.FILES.get('file')
    if not f:
        return Response({'message': 'No file uploaded'}, status=400)
    xlsx_buf = f.read()
    try:
        pdf_buf = email_svc.convert_excel_to_pdf(
            xlsx_buf, print_area='$A$1:$AD$53', setup_page=True,
        )
    except Exception as e:
        return Response({'message': f'PDF conversion failed: {e}'}, status=500)

    province = sanitize_fs_name(request.data.get('province') or 'Unknown Province')
    abattoir = sanitize_fs_name(request.data.get('abattoir') or 'Unknown Abattoir')
    training_date_raw = request.data.get('training_date') or ''
    # Parse training start date (YYYY-MM-DD from frontend date input)
    try:
        td = datetime.strptime(training_date_raw, '%Y-%m-%d')
    except (ValueError, TypeError):
        td = datetime.now()
    now = datetime.now()
    folder_name = (
        f"STT Training {td.day:02d}-{td.month:02d}-{td.year} "
        f"{now.hour:02d}-{now.minute:02d}-{now.second:02d} {abattoir}"
    )
    save_dir = settings.DOCUMENTS_ROOT / province / abattoir / 'STT Training Documents' / folder_name
    save_dir.mkdir(parents=True, exist_ok=True)
    (save_dir / f'{folder_name}.xlsx').write_bytes(xlsx_buf)
    (save_dir / f'{folder_name}.pdf').write_bytes(pdf_buf)
    return Response({'ok': True})


@api_view(['GET'])
def stt_breakdown_view(request):
    vendor = connection.vendor
    yfn = ((lambda c: f"CAST(strftime('%Y', {c}) AS INTEGER)")
           if vendor == 'sqlite' else (lambda c: f'YEAR({c})'))
    mfn = ((lambda c: f"CAST(strftime('%m', {c}) AS INTEGER)")
           if vendor == 'sqlite' else (lambda c: f'MONTH({c})'))

    where, params = [], []
    year = request.query_params.get('year') or ''
    month = request.query_params.get('month') or ''
    quarter = request.query_params.get('quarter') or ''
    province = request.query_params.get('province') or ''
    abattoir = request.query_params.get('abattoir') or ''

    if year:
        yrs = [int(y) for y in year.split(',') if y.strip().isdigit()]
        yrs = [y for y in yrs if 0 < y < 3000]
        if yrs:
            ph = ','.join(['%s'] * len(yrs))
            where.append(f"{yfn('training_start_date')} IN ({ph})")
            params.extend(yrs)
    if month:
        mos = [int(m) for m in month.split(',') if m.strip().isdigit()]
        mos = [m for m in mos if 1 <= m <= 12]
        if mos:
            ph = ','.join(['%s'] * len(mos))
            where.append(f"{mfn('training_start_date')} IN ({ph})")
            params.extend(mos)
    if quarter:
        try:
            q = int(quarter)
            where.append(f"{mfn('training_start_date')} BETWEEN %s AND %s")
            params.extend([(q - 1) * 3 + 1, q * 3])
        except ValueError:
            pass
    if province:
        provs = [p for p in province.split(',') if p]
        if provs:
            ph = ','.join(['%s'] * len(provs))
            where.append(f"province IN ({ph})")
            params.extend(provs)
    if abattoir:
        abs_list = [a for a in abattoir.split(',') if a]
        if len(abs_list) == 1:
            where.append("abattoir_name LIKE %s")
            params.append(f'%{abs_list[0]}%')
        elif len(abs_list) > 1:
            ph = ','.join(['%s'] * len(abs_list))
            where.append(f"abattoir_name IN ({ph})")
            params.extend(abs_list)

    where_sql = ('WHERE ' + ' AND '.join(where)) if where else ''
    if_ref = safe_col_ref('if_', {'if_'})

    sql = f"""
        SELECT
          training_start_date, training_end_date, abattoir_name,
          municipality, province, thru_put, specie,
          COUNT(*) AS total_trained,
          SUM(CASE WHEN am   = '1' THEN 1 ELSE 0 END) AS am,
          SUM(CASE WHEN af   = '1' THEN 1 ELSE 0 END) AS af,
          SUM(CASE WHEN cm   = '1' THEN 1 ELSE 0 END) AS cm,
          SUM(CASE WHEN cf   = '1' THEN 1 ELSE 0 END) AS cf,
          SUM(CASE WHEN im   = '1' THEN 1 ELSE 0 END) AS im,
          SUM(CASE WHEN {if_ref} = '1' THEN 1 ELSE 0 END) AS if_,
          SUM(CASE WHEN wm   = '1' THEN 1 ELSE 0 END) AS wm,
          SUM(CASE WHEN wf   = '1' THEN 1 ELSE 0 END) AS wf,
          SUM(CASE WHEN age_lt35  = '1' THEN 1 ELSE 0 END) AS age_lt35,
          SUM(CASE WHEN age_35_55 = '1' THEN 1 ELSE 0 END) AS age_35_55,
          SUM(CASE WHEN age_gt55  = '1' THEN 1 ELSE 0 END) AS age_gt55,
          SUM(CASE WHEN am='1' OR af='1' OR cm='1' OR cf='1' OR im='1' OR {if_ref}='1' THEN 1 ELSE 0 END) AS hdis,
          SUM(CASE WHEN disability = 'Yes' THEN 1 ELSE 0 END) AS disability_count
        FROM STTTrainingReport
        {where_sql}
        GROUP BY training_start_date, training_end_date, abattoir_name,
                 municipality, province, thru_put, specie
        ORDER BY training_start_date ASC, abattoir_name ASC, specie ASC
    """

    with connection.cursor() as c:
        c.execute(sql, params)
        rows = rows_to_dicts(c)
        c.execute(f"""
            SELECT DISTINCT {yfn('training_start_date')} AS yr, province, abattoir_name
            FROM STTTrainingReport
            WHERE training_start_date IS NOT NULL AND training_start_date <> ''
        """)
        filt = rows_to_dicts(c)
    years_list = sorted({r['yr'] for r in filt if r['yr']}, reverse=True)
    provinces_list = sorted({r['province'] for r in filt if r['province']})
    abattoirs_list = sorted({r['abattoir_name'] for r in filt if r['abattoir_name']})
    return Response({
        'rows': rows, 'years': years_list,
        'provinces': provinces_list, 'abattoirs': abattoirs_list,
    })


# ===========================================================================
#  Residue monitoring
# ===========================================================================
@api_view(['GET'])
def residue_template_view(request):
    path = Path(settings.PROJECT_ROOT) / 'Residue Monitoring Upload Template.xlsx'
    if not path.exists():
        return Response({'message': 'Template file not found.'}, status=500)
    return FileResponse(
        path.open('rb'),
        as_attachment=True,
        filename='Residue Monitoring Upload Template.xlsx',
    )


@api_view(['POST'])
@parser_classes([MultiPartParser, FormParser])
def residue_upload_view(request):
    from openpyxl import load_workbook
    f = request.FILES.get('file')
    if not f:
        return Response({'message': 'No file uploaded.'}, status=400)
    try:
        wb = load_workbook(BytesIO(f.read()), data_only=True)
        batch_id = str(uuid.uuid4())
        all_rows = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            species = sheet_name.strip().split(' ')[-1]
            rows_iter = list(ws.iter_rows(values_only=True))
            for i in range(1, len(rows_iter)):
                row = rows_iter[i]
                if not row or all(v in (None, '') for v in row):
                    continue
                obj = {'species': species, 'batch_id': batch_id}
                for idx, key in enumerate(RESIDUE_HEADER_MAP):
                    val = row[idx] if idx < len(row) else ''
                    obj[key] = str(val) if val is not None else ''
                raw = f"{obj.get('sample_ref','')}|{obj.get('date_collected','')}|{obj.get('sample_type','')}|{species}"
                obj['record_id'] = hashlib.sha256(raw.encode()).hexdigest()
                all_rows.append(obj)

        col_list = ['batch_id'] + RESIDUE_INSERT_COLS
        batch_n = 40
        with connection.cursor() as c:
            for i in range(0, len(all_rows), batch_n):
                chunk = all_rows[i:i + batch_n]
                placeholders = ','.join(
                    '(' + ','.join(['%s'] * len(col_list)) + ')' for _ in chunk
                )
                flat_vals: list = []
                for r in chunk:
                    for col in col_list:
                        flat_vals.append(r.get(col, '') or '')
                c.execute(
                    f"INSERT INTO ResidueMonitoringTemp ({','.join(col_list)}) VALUES {placeholders}",
                    flat_vals,
                )
        return Response({'batchId': batch_id, 'rowCount': len(all_rows)})
    except Exception as e:
        return Response({'message': f'Failed to process file: {e}'}, status=500)


@api_view(['GET', 'DELETE'])
def residue_temp_view(request, batch_id):
    if request.method == 'DELETE':
        with connection.cursor() as c:
            c.execute('DELETE FROM ResidueMonitoringTemp WHERE batch_id = %s', [batch_id])
        return Response({'message': 'Batch discarded.'})

    with connection.cursor() as c:
        c.execute(
            'SELECT * FROM ResidueMonitoringTemp WHERE batch_id = %s ORDER BY id',
            [batch_id],
        )
        rows = rows_to_dicts(c)
    return Response(rows)


@api_view(['POST'])
def residue_commit_view(request):
    batch_id = request.data.get('batchId')
    rows = request.data.get('rows')
    if not batch_id or not isinstance(rows, list):
        return Response({'message': 'Invalid request.'}, status=400)

    with connection.cursor() as c:
        c.execute("SELECT sample_ref, date_collected, sample_type, species FROM ResidueMonitoring")
        existing = set()
        for r in c.fetchall():
            existing.add('|'.join((r[i] or '').strip() for i in range(4)))

    new_rows = []
    dup_count = 0
    for row in rows:
        key = '|'.join(str(row.get(k, '') or '').strip()
                       for k in ('sample_ref', 'date_collected', 'sample_type', 'species'))
        if key in existing:
            dup_count += 1
        else:
            new_rows.append(row)
            existing.add(key)

    with connection.cursor() as c:
        if new_rows:
            batch_n = 40
            for i in range(0, len(new_rows), batch_n):
                chunk = new_rows[i:i + batch_n]
                placeholders = ','.join(
                    '(' + ','.join(['%s'] * len(RESIDUE_INSERT_COLS)) + ')' for _ in chunk
                )
                flat_vals: list = []
                for r in chunk:
                    for col in RESIDUE_INSERT_COLS:
                        flat_vals.append(r.get(col, '') or '')
                c.execute(
                    f"INSERT INTO ResidueMonitoring ({','.join(RESIDUE_INSERT_COLS)}) "
                    f"VALUES {placeholders}",
                    flat_vals,
                )
        c.execute('DELETE FROM ResidueMonitoringTemp WHERE batch_id = %s', [batch_id])

    msg = f'{len(new_rows)} records committed.'
    if dup_count:
        msg += f' {dup_count} duplicate(s) skipped.'
    return Response({'message': msg, 'committed': len(new_rows), 'duplicates': dup_count})


@api_view(['GET'])
def residue_committed_view(request):
    try:
        page = max(1, int(request.query_params.get('page') or '1'))
        page_size = min(99999, int(request.query_params.get('size') or '200'))
    except ValueError:
        page, page_size = 1, 200
    offset = (page - 1) * page_size

    reserved = {'page', 'size'}
    where_parts, params = apply_column_filters(
        request.query_params.dict(), RESIDUE_ALL_COLS, set(), reserved,
    )
    where_sql = ('WHERE ' + ' AND '.join(where_parts)) if where_parts else ''

    with connection.cursor() as c:
        c.execute(f'SELECT COUNT(*) FROM ResidueMonitoring {where_sql}', params)
        total = c.fetchone()[0]
        c.execute(
            f'SELECT * FROM ResidueMonitoring {where_sql} ORDER BY id LIMIT %s OFFSET %s',
            params + [page_size, offset],
        )
        rows = rows_to_dicts(c)
    return Response({'total': total, 'page': page, 'pageSize': page_size, 'rows': rows})


@api_view(['PUT', 'PATCH'])
def residue_committed_update_view(request, pk):
    row = request.data or {}
    sets = ', '.join(f'{c} = %s' for c in RESIDUE_ALL_COLS)
    values = [str(row.get(c, '') or '') for c in RESIDUE_ALL_COLS] + [pk]
    with connection.cursor() as c:
        c.execute(f'UPDATE ResidueMonitoring SET {sets} WHERE id = %s', values)
    return Response({'ok': True})


@api_view(['DELETE'])
def residue_clear_all_view(request):
    with connection.cursor() as c:
        c.execute('DELETE FROM ResidueMonitoring')
    return Response({'ok': True})


# ===========================================================================
#  Quotation
# ===========================================================================
@api_view(['GET'])
def quotation_abattoir_details_view(request):
    name = request.query_params.get('name')
    if not name:
        return Response({'message': 'name required'}, status=400)
    year = datetime.now().year
    member_col = f'member_{year}'
    if member_col not in ABATTOIR_COLS:
        member_col = 'member_2025'
    with connection.cursor() as c:
        c.execute(
            f"""SELECT abattoir_name, rc_nr, lh, vat_number, province, municipality,
                       {member_col} AS is_member
                FROM AbattoirMaster WHERE abattoir_name = %s""",
            [name],
        )
        rows = rows_to_dicts(c)
    if not rows:
        return Response({'found': False})
    row = rows[0]
    is_member = 'Yes' if 'member' in (row.get('is_member') or '').lower() else 'No'
    return Response({
        'found': True,
        'abattoir_name': row['abattoir_name'],
        'rc_nr': row.get('rc_nr') or '',
        'thru_put': row.get('lh') or '',
        'vat_number': row.get('vat_number') or '',
        'province': row.get('province') or '',
        'municipality': row.get('municipality') or '',
        'is_member': is_member,
    })


@api_view(['POST'])
def quotation_generate_view(request):
    data = request.data or {}
    template_path = Path(settings.PROJECT_ROOT) / 'Quotation Template.xlsx'
    if not template_path.exists():
        return Response({'message': 'Template not found'}, status=500)

    try:
        with open(template_path, 'rb') as f:
            template_bytes = f.read()

        today = datetime.now()
        date_str = today.strftime('%d/%m/%Y')

        def safe_float(v):
            try: return float(v) if v else None
            except (ValueError, TypeError): return None

        cell_updates = {
            'D2': f"Training and Support Services Contract {today.year}",
            'A2': f"QUOTATION DATE: {date_str}",
            'D3': data.get('clientName') or '',
            'D4': data.get('rmaaMember') or '',
            'D5': data.get('rc') or '',
            'D6': data.get('throughput') or '',
            'D7': data.get('vatNumber') or '',
            'D9': data.get('clientContact') or '',
            'D10': data.get('telephone') or '',
            'D12': data.get('cell') or '',
            'D13': data.get('email') or '',
            'D14': data.get('postalAddress') or '',
            'D15': data.get('streetAddress') or '',
            'D16': data.get('rmaaContact') or '',
        }

        # Line items (rows 22-26, up to 5)
        for i, item in enumerate(data.get('lineItems') or []):
            row = 22 + i
            if row > 26:
                break
            try:
                d = datetime.fromisoformat(item.get('date')) if item.get('date') else None
                cell_updates[f'B{row}'] = d.strftime('%A, %B %d, %Y') if d else ''
            except (ValueError, TypeError):
                cell_updates[f'B{row}'] = item.get('date') or ''
            skill = item.get('skillsProgramme') or ''
            qty = item.get('qty') or ''
            cell_updates[f'C{row}'] = f"{skill} x {qty}" if skill and qty else skill
            pc = safe_float(item.get('programmeCost'))
            pq = int(item.get('qty') or 1) if item.get('qty') else 1
            cell_updates[f'D{row}'] = pc * pq if pc is not None else ''
            slaught = item.get('slaughterTechnique') or ''
            sq = item.get('slaughterQty') or ''
            cell_updates[f'E{row}'] = f"{slaught} x {sq}" if slaught and sq else slaught
            sc = safe_float(item.get('slaughterCost'))
            sqn = int(item.get('slaughterQty') or 1) if item.get('slaughterQty') else 1
            cell_updates[f'F{row}'] = sc * sqn if sc is not None else ''
            cell_updates[f'G{row}'] = safe_float(item.get('distance')) or ''
            cell_updates[f'H{row}'] = safe_float(item.get('accommodation')) or ''

        # Sampling (row 29)
        samp = data.get('sampling') or {}
        samp_qty = samp.get('qty') or ''
        samp_cost = safe_float(samp.get('cost'))
        if samp_qty:
            cell_updates['B29'] = f"Sampling x {samp_qty}"
            cell_updates['F29'] = samp_cost * int(samp_qty) if samp_cost is not None else ''
            cell_updates['G29'] = safe_float(samp.get('distance')) or ''
            cell_updates['H29'] = safe_float(samp.get('accommodation')) or ''

        # Audit Verification (row 30)
        aud = data.get('audit') or {}
        aud_qty = aud.get('qty') or ''
        aud_cost = safe_float(aud.get('cost'))
        if aud_qty:
            cell_updates['B30'] = f"Verification Audit x {aud_qty}"
            cell_updates['F30'] = aud_cost * int(aud_qty) if aud_cost is not None else ''
            cell_updates['G30'] = safe_float(aud.get('distance')) or ''
            cell_updates['H30'] = safe_float(aud.get('accommodation')) or ''

        # Discount lines (rows 34-37)
        disc = data.get('discounts')
        if disc:
            disc_rows = [
                (34, 'Skills Program', 'skillsAmount', 'skillsKm', 'skillsAccomm'),
                (35, 'Sampling', 'samplingAmount', 'samplingKm', 'samplingAccomm'),
                (36, 'Verification Audit', 'auditAmount', 'auditKm', 'auditAccomm'),
                (37, 'Membership', 'membershipAmount', 'membershipKm', 'membershipAccomm'),
            ]
            for row, label, amt_key, km_key, acc_key in disc_rows:
                amt = safe_float(disc.get(amt_key))
                km = safe_float(disc.get(km_key))
                acc = safe_float(disc.get(acc_key))
                if amt or km or acc:
                    cell_updates[f'D{row}'] = label
                    cell_updates[f'F{row}'] = -abs(amt) if amt is not None else ''
                    cell_updates[f'G{row}'] = -abs(km) if km is not None else ''
                    cell_updates[f'H{row}'] = -abs(acc) if acc is not None else ''

        xlsx_bytes = email_svc.modify_xlsx_cells(template_bytes, cell_updates)
        pdf_bytes = email_svc.convert_excel_to_pdf(xlsx_bytes, print_area='$A$1:$I$45', setup_page=True)
        folder = f"Quotation {date_str.replace('/', '-')} {sanitize_fs_name(data.get('clientName'))}"
        return Response({
            'ok': True,
            'pdfBase64': base64.b64encode(pdf_bytes).decode('ascii'),
            'xlsxBase64': base64.b64encode(xlsx_bytes).decode('ascii'),
            'fileName': f'{folder}.pdf',
            'folderName': folder,
            'province': data.get('province') or '',
            'clientName': data.get('clientName') or '',
        })
    except Exception as e:
        return Response({'message': str(e)}, status=500)


@api_view(['POST'])
def quotation_send_view(request):
    data = request.data or {}
    to = data.get('to')
    pdf_b64 = data.get('pdfBase64')
    if not to or not pdf_b64:
        return Response({'message': 'Missing email or PDF'}, status=400)
    cc_list = data.get('cc') or []
    result = email_svc.send_quotation_email(
        to=to, cc=cc_list, client_name=data.get('clientName') or '',
        pdf_base64=pdf_b64, file_name=data.get('fileName') or 'Quotation.pdf',
    )
    if not result['ok']:
        return Response({'message': f"Email failed: {result.get('reason')}"}, status=500)

    folder = data.get('folderName')
    province = data.get('province')
    client = data.get('clientName')
    xlsx_b64 = data.get('xlsxBase64')
    if folder and province and client:
        try:
            save_dir = (
                settings.DOCUMENTS_ROOT
                / sanitize_fs_name(province)
                / sanitize_fs_name(client)
                / 'Quotations' / folder
            )
            save_dir.mkdir(parents=True, exist_ok=True)
            if xlsx_b64:
                (save_dir / f'{folder}.xlsx').write_bytes(base64.b64decode(xlsx_b64))
            (save_dir / f'{folder}.pdf').write_bytes(base64.b64decode(pdf_b64))
        except Exception:
            pass
    return Response({'ok': True})


# ===========================================================================
#  Documents
# ===========================================================================
def _safe_abs(rel: str) -> Path:
    abs_path = (settings.DOCUMENTS_ROOT / rel).resolve()
    root_resolved = settings.DOCUMENTS_ROOT.resolve()
    if not str(abs_path).startswith(str(root_resolved)):
        raise ValueError('Invalid path')
    return abs_path


def _read_tree(dir_path: Path, rel_base: str = '') -> list:
    out_folders = []
    out_files = []
    for e in sorted(dir_path.iterdir(), key=lambda p: p.name.lower()):
        rel = f'{rel_base}/{e.name}' if rel_base else e.name
        if e.is_dir():
            children = _read_tree(e, rel)
            if children:
                out_folders.append({
                    'type': 'folder', 'name': e.name, 'path': rel, 'children': children,
                })
        elif e.is_file() and re.search(r'\.(pdf|xlsx)$', e.name, re.I):
            st = e.stat()
            out_files.append({
                'type': 'file', 'name': e.name, 'path': rel,
                'size': st.st_size, 'modified': st.st_mtime,
            })
    return out_folders + out_files


def _purge_empty(dir_path: Path, root: Path):
    for child in list(dir_path.iterdir()):
        if child.is_dir():
            _purge_empty(child, root)
    if dir_path != root and not any(dir_path.iterdir()):
        dir_path.rmdir()


@api_view(['GET'])
def documents_tree_view(request):
    root = settings.DOCUMENTS_ROOT
    root.mkdir(parents=True, exist_ok=True)
    try:
        _purge_empty(root, root)
    except OSError:
        pass
    return Response({'tree': _read_tree(root, '')})


@api_view(['GET'])
def documents_view_view(request):
    rel = request.query_params.get('p')
    if not rel:
        return Response({'message': 'Missing path'}, status=400)
    try:
        abs_path = _safe_abs(rel)
    except ValueError:
        return Response({'message': 'Invalid path'}, status=400)
    if not abs_path.exists():
        return Response({'message': 'File not found'}, status=404)
    is_xlsx = str(abs_path).lower().endswith('.xlsx')
    content_type = (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        if is_xlsx else 'application/pdf'
    )
    resp = FileResponse(abs_path.open('rb'), content_type=content_type)
    if is_xlsx:
        resp['Content-Disposition'] = f'attachment; filename="{abs_path.name}"'
    else:
        resp['Content-Disposition'] = 'inline'
    resp['X-Frame-Options'] = 'SAMEORIGIN'
    return resp


@api_view(['GET'])
def documents_download_view(request):
    rel = request.query_params.get('p')
    if not rel:
        return Response({'message': 'Missing path'}, status=400)
    try:
        abs_path = _safe_abs(rel)
    except ValueError:
        return Response({'message': 'Invalid path'}, status=400)
    if not abs_path.exists():
        return Response({'message': 'File not found'}, status=404)
    return FileResponse(abs_path.open('rb'), as_attachment=True, filename=abs_path.name)


@api_view(['DELETE'])
def documents_delete_view(request):
    rel = request.query_params.get('p')
    if not rel:
        return Response({'message': 'Missing path'}, status=400)
    try:
        abs_path = _safe_abs(rel)
    except ValueError:
        return Response({'message': 'Invalid path'}, status=400)
    if not abs_path.exists():
        return Response({'message': 'File not found'}, status=404)
    abs_path.unlink()
    parent = abs_path.parent
    root = settings.DOCUMENTS_ROOT.resolve()
    while parent != root and str(parent).startswith(str(root)):
        try:
            parent.rmdir()
        except OSError:
            break
        parent = parent.parent
    return Response({'ok': True})


# ===========================================================================
#  facilitators
# ===========================================================================
@api_view(['GET', 'POST'])
def facilitators_view(request):
    if request.method == 'POST':
        name = (request.data.get('name') or '').strip()
        surname = (request.data.get('surname') or '').strip()
        if not name:
            return Response({'message': 'Name is required.'}, status=400)
        if Facilitator.objects.filter(name=name, surname=surname).exists():
            return Response({'message': 'Facilitator already exists.'}, status=400)
        f = Facilitator.objects.create(name=name, surname=surname)
        return Response({'ok': True, 'id': f.id, 'name': f.name, 'surname': f.surname})
    # GET — list all
    facilitators = list(
        Facilitator.objects.order_by('name', 'surname').values('id', 'name', 'surname')
    )
    return Response({'facilitators': facilitators})


@api_view(['DELETE'])
def facilitator_delete_view(request, pk):
    try:
        Facilitator.objects.filter(id=pk).delete()
    except Exception as e:
        return Response({'message': str(e)}, status=500)
    return Response({'ok': True})
