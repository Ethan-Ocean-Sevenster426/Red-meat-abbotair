"""Load seed data from seed-data/*.xlsx into MySQL.

Each xlsx filename maps to its DB table (e.g. AbattoirMaster.xlsx → AbattoirMaster).
The first row is the header whose values are DB column names.
Preserves primary keys so foreign references (AuditLog.record_id) stay valid.

Run: ./manage.py seed_from_xlsx [--truncate]
"""
from pathlib import Path
from django.conf import settings
from django.core.management.base import BaseCommand
from django.db import connection
from openpyxl import load_workbook


# filename (no .xlsx) → (table name, columns to skip when inserting)
FILE_TO_TABLE = {
    'Users': 'Users',
    'AbattoirMaster': 'AbattoirMaster',
    'AssociatedMembersMaster': 'AssociatedMembersMaster',
    'AuditLog': 'AuditLog',
    'GovernmentMaster': 'GovernmentMaster',
    'IndustryMaster': 'IndustryMaster',
    'STTTrainingReport': 'STTTrainingReport',
    'TransformationMaster': 'TransformationMaster',
    'UserColumnPreferences': 'UserColumnPreferences',
}


class Command(BaseCommand):
    help = 'Import seed-data/*.xlsx into the DB, preserving IDs.'

    def add_arguments(self, parser):
        parser.add_argument('--truncate', action='store_true',
                            help='Delete existing rows before inserting.')
        parser.add_argument('--dir', default=None, help='Override seed-data directory.')

    def handle(self, *args, **options):
        seed_dir = Path(options['dir']) if options['dir'] else (
            Path(settings.PROJECT_ROOT) / 'seed-data'
        )
        if not seed_dir.exists():
            self.stderr.write(f'No seed directory: {seed_dir}')
            return

        vendor = connection.vendor
        quote = '`' if vendor == 'mysql' else '"'

        # Seed Users first (others reference nothing we'd block on), AuditLog last
        # to let the referenced records exist.
        priority = ['Users', 'AbattoirMaster', 'TransformationMaster', 'GovernmentMaster',
                    'IndustryMaster', 'AssociatedMembersMaster', 'STTTrainingReport',
                    'UserColumnPreferences', 'AuditLog']

        # Temporarily disable FK checks for MySQL to allow flexible ordering
        if vendor == 'mysql':
            with connection.cursor() as c:
                c.execute('SET FOREIGN_KEY_CHECKS=0')

        try:
            for name in priority:
                table = FILE_TO_TABLE.get(name)
                if not table:
                    continue
                path = seed_dir / f'{name}.xlsx'
                if not path.exists():
                    self.stdout.write(f'  skip (no file): {path.name}')
                    continue
                self._import_file(path, table, options['truncate'], quote)
        finally:
            if vendor == 'mysql':
                with connection.cursor() as c:
                    c.execute('SET FOREIGN_KEY_CHECKS=1')

        self.stdout.write(self.style.SUCCESS('Done.'))

    def _import_file(self, path: Path, table: str, truncate: bool, quote: str):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]

        rows = ws.iter_rows(values_only=True)
        try:
            headers = next(rows)
        except StopIteration:
            self.stdout.write(f'  {path.name}: empty')
            return
        headers = [h for h in headers if h is not None]
        n_cols = len(headers)
        col_list = ','.join(f'{quote}{h}{quote}' for h in headers)
        placeholders = ','.join(['%s'] * n_cols)

        if truncate:
            with connection.cursor() as c:
                c.execute(f'DELETE FROM {table}')
                if connection.vendor == 'mysql':
                    c.execute(f'ALTER TABLE {table} AUTO_INCREMENT = 1')

        batch, count = [], 0
        BATCH_SIZE = 500
        sql = f'INSERT INTO {table} ({col_list}) VALUES ({placeholders})'

        with connection.cursor() as c:
            for row in rows:
                vals = [self._coerce(v) for v in row[:n_cols]]
                # pad short rows
                if len(vals) < n_cols:
                    vals += [None] * (n_cols - len(vals))
                batch.append(vals)
                if len(batch) >= BATCH_SIZE:
                    c.executemany(sql, batch)
                    count += len(batch)
                    batch = []
            if batch:
                c.executemany(sql, batch)
                count += len(batch)

        self.stdout.write(f'  {path.name}: inserted {count} rows')

    @staticmethod
    def _coerce(v):
        """Convert openpyxl cell value to a DB-friendly python value.

        Empty strings become NULL (MySQL DATETIME rejects '').
        ISO datetime strings ('2026-04-14T09:57:52.406Z') are parsed to
        datetime objects so MySQL accepts them.
        """
        from datetime import datetime
        import re
        if v is None:
            return None
        if isinstance(v, str):
            if v == '':
                return None
            if re.match(r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}', v):
                try:
                    return datetime.fromisoformat(v.replace('Z', '+00:00'))
                except ValueError:
                    return v
            return v
        return str(v)
